"""
GeM Scraper — Web Dashboard
Flask-based web interface for viewing scraped tender data, PDFs, and summaries.
"""
import json
import os
import mimetypes as mtypes
import markdown
import subprocess
from pathlib import Path
from flask import Flask, render_template, send_file, jsonify, request, abort, redirect, url_for

import config
import db

app = Flask(
    __name__,
    static_folder="static_web",
    static_url_path="/static",
    template_folder="templates_web",
)
app.config["SECRET_KEY"] = "gem-scraper-dashboard"


def bid_to_folder(bid_number: str) -> str:
    """Convert bid number to folder-safe name: GEM/2026/B/123 -> GEM-2026-B-123"""
    return bid_number.replace("/", "-")


def folder_to_bid(folder: str) -> str:
    """Convert folder name back to bid number: GEM-2026-B-123 -> GEM/2026/B/123"""
    parts = folder.split("-")
    if len(parts) >= 4 and parts[0] == "GEM":
        return f"{parts[0]}/{parts[1]}/{parts[2]}/{'-'.join(parts[3:])}"
    return folder.replace("-", "/")


def detect_file_type(file_path: Path) -> str:
    """Detect file type by extension or magic bytes."""
    ext = file_path.suffix.lower()
    if ext in (".pdf", ".xlsx", ".xls", ".csv", ".doc", ".docx", ".zip"):
        return ext

    # No extension — detect by reading first bytes
    try:
        with open(file_path, "rb") as f:
            header = f.read(8)
        if header[:5] == b"%PDF-":
            return ".pdf"
        if header[:4] == b"PK\x03\x04":  # ZIP header (xlsx/docx are ZIP)
            return ".xlsx"
        if header[:2] == b"\xd0\xcf":  # OLE2 (xls/doc)
            return ".xls"
    except Exception:
        pass
    return ext or ".bin"


def get_attachments(bid_folder: str) -> list[dict]:
    """Get all valid attachments for a tender with type detection."""
    folder = config.TENDERS_DIR / bid_folder / "attachments"
    attachments = []
    if not folder.exists():
        return attachments

    for f in sorted(folder.iterdir()):
        if not f.is_file() or f.stat().st_size == 0:
            continue
        file_type = detect_file_type(f)
        attachments.append({
            "name": f.name,
            "size": f.stat().st_size,
            "ext": file_type,
            "previewable": file_type in (".pdf", ".xlsx", ".xls", ".csv"),
            "is_pdf": file_type == ".pdf",
            "is_spreadsheet": file_type in (".xlsx", ".xls", ".csv"),
        })
    return attachments


@app.route("/")
def dashboard():
    """Main dashboard — list all tenders."""
    stats = db.get_tender_stats()
    tenders = db.get_all_tenders(limit=200)
    # Add folder name for URL building
    locations = set()
    for t in tenders:
        t["folder"] = bid_to_folder(t["bid_number"])
        loc = t.get("location") or ""
        if loc and loc != "null" and loc != "None" and loc.strip():
            locations.add(loc.strip())
    # Sort locations alphabetically
    locations = sorted(locations)
    return render_template("dashboard.html", tenders=tenders, stats=stats, locations=locations)


@app.route("/tender/<bid_folder>")
def tender_detail(bid_folder):
    """Detail view for a single tender."""
    bid_number = folder_to_bid(bid_folder)
    tender = db.get_tender_by_bid(bid_number)
    if not tender:
        abort(404)

    links = db.get_links_for_tender(tender["id"])
    relevant_links = [l for l in links if l.get("is_relevant")]
    other_links = [l for l in links if not l.get("is_relevant")]

    # Read summary markdown
    summary_md = ""
    folder = config.TENDERS_DIR / bid_folder
    summary_file = folder / "summary.md"
    if summary_file.exists():
        raw_md = summary_file.read_text(encoding="utf-8")
        summary_md = markdown.markdown(raw_md, extensions=["tables", "fenced_code"])

    # Get attachments with preview capability
    attachments = get_attachments(bid_folder)

    # Check if tender PDF exists
    has_pdf = (folder / "tender.pdf").exists()

    return render_template(
        "tender_detail.html",
        tender=tender,
        all_links=links,
        relevant_links=relevant_links,
        other_links=other_links,
        summary_html=summary_md,
        attachments=attachments,
        has_pdf=has_pdf,
        bid_folder=bid_folder,
    )


@app.route("/pdf/<bid_folder>")
def view_pdf(bid_folder):
    """Serve a tender PDF for inline viewing."""
    pdf_path = config.TENDERS_DIR / bid_folder / "tender.pdf"
    if not pdf_path.exists():
        abort(404)
    return send_file(str(pdf_path), mimetype="application/pdf")


@app.route("/attachment/<bid_folder>/<filename>")
def view_attachment(bid_folder, filename):
    """Serve an attachment file (inline for PDFs/spreadsheets)."""
    file_path = config.TENDERS_DIR / bid_folder / "attachments" / filename
    if not file_path.exists():
        abort(404)

    file_type = detect_file_type(file_path)
    type_map = {
        ".pdf": "application/pdf",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xls": "application/vnd.ms-excel",
        ".csv": "text/csv",
        ".doc": "application/msword",
        ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ".zip": "application/zip",
    }
    mimetype = type_map.get(file_type, "application/octet-stream")
    return send_file(str(file_path), mimetype=mimetype)


@app.route("/preview/xlsx/<bid_folder>/<filename>")
def preview_xlsx(bid_folder, filename):
    """Render an XLSX/XLS/CSV file as an HTML table for inline preview."""
    file_path = config.TENDERS_DIR / bid_folder / "attachments" / filename
    if not file_path.exists():
        abort(404)

    file_type = detect_file_type(file_path)
    try:
        if file_type == ".csv":
            import csv
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                rows = list(reader)
            if not rows:
                return "<p>Empty spreadsheet</p>"
            headers = rows[0]
            data_rows = rows[1:200]  # limit to 200 rows
        else:
            import openpyxl
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
            sheets_html = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i > 200:
                        break
                    rows.append([str(c) if c is not None else "" for c in row])
                if not rows:
                    continue
                headers = rows[0]
                data_rows = rows[1:]
                sheets_html.append((sheet_name, headers, data_rows))
            wb.close()

            if not sheets_html:
                return "<p>Empty spreadsheet</p>"

            # Render all sheets
            html = ""
            for sheet_name, headers, data_rows in sheets_html:
                html += f'<h3 style="color:#a78bfa;margin:1rem 0 0.5rem;font-size:0.9rem;">Sheet: {sheet_name}</h3>'
                html += '<div style="overflow-x:auto;"><table class="xlsx-table"><thead><tr>'
                for h in headers:
                    html += f"<th>{h}</th>"
                html += "</tr></thead><tbody>"
                for row in data_rows:
                    html += "<tr>"
                    for cell in row:
                        html += f"<td>{cell}</td>"
                    html += "</tr>"
                html += "</tbody></table></div>"
            return html

        # CSV rendering
        html = '<div style="overflow-x:auto;"><table class="xlsx-table"><thead><tr>'
        for h in headers:
            html += f"<th>{h}</th>"
        html += "</tr></thead><tbody>"
        for row in data_rows:
            html += "<tr>"
            for cell in row:
                html += f"<td>{cell}</td>"
            html += "</tr>"
        html += "</tbody></table></div>"
        return html

    except Exception as e:
        return f'<p style="color:#ef4444;">Failed to preview: {str(e)}</p>'


@app.route("/api/stats")
def api_stats():
    return jsonify(db.get_tender_stats())


@app.route("/api/tenders")
def api_tenders():
    tenders = db.get_all_tenders(limit=200)
    return jsonify(tenders)


@app.route("/api/tender/<bid_folder>")
def api_tender(bid_folder):
    bid_number = folder_to_bid(bid_folder)
    tender = db.get_tender_by_bid(bid_number)
    if not tender:
        abort(404)
    links = db.get_links_for_tender(tender["id"])
    return jsonify({"tender": tender, "links": links})


if __name__ == "__main__":
    print("\n🌐 GeM Scraper Dashboard")
    print("   http://localhost:5050\n")
    app.run(host="0.0.0.0", port=5050, debug=True)
